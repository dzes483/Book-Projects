#!/usr/bin/python3
# text_files_to_spreadsheet.py - Reads the contents of the desired spreadsheet
# and pastes the data into a separate text file, per column of data.

import openpyxl

os.chdir('C:\\Users\\Jessica\\AppData\\Local\\Programs\\Python\\Python36\\MyPythonScripts\\Chapter 12')
filename = 'test_file'
# Load workbook and define sheet
file = str('C:\\Users\\Jessica\\AppData\\Local\\Programs\\Python\\Python36\\MyPythonScripts\\My Chapter Projects\\Chapter 12\\test_file.xlsx')
wb = openpyxl.load_workbook(file)
sheet = wb.active

col_counter = 1

for cols in range(1, sheet.max_column + 1):
    data = []
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=col_counter).value != None:
            data.append(sheet.cell(row=row, column=col_counter).value)

    new_file = open(filename + '_' + str(col_counter) + '.txt', 'w')
    for i in data:
        new_file.write(i + '\n')
    new_file.close()

    col_counter += 1
