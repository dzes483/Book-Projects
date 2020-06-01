#!/usr/bin/python3
# multiplication_table.py - Takes a number N from the user and creates
# an NxN multiplication table in an Excel spreadsheet.

import openpyxl
from openpyxl.styles import Font

# Create the workbook and define the worksheet
wb = openpyxl.Workbook()
sheet = wb.active

# Get the number from the user
while True:
    try:
        num = int(input('Input a number: '))
    except ValueError:
        print('Invalid character. Please input a number.')
        continue
    else:
        break

# Define the counters and the font style
bold_text = Font(bold=True)
row_counter = 1
col_counter = 1
column_num = 2

# Fill out the rows with the header numbers
print('Printing headers...')
for row in range(2, num+2):
    sheet['A' + str(row)].value = row_counter
    sheet['A' + str(row)].font = bold_text
    row_counter += 1
# Fill out the columns with the header numbers
for col in range(1, num+1):
    sheet.cell(row=1, column=column_num).value = col_counter
    sheet.cell(row=1, column=column_num).font = bold_text
    col_counter += 1
    column_num += 1

# Fill out the data
print('Filling out the data...')
for cols in sheet:
    for cell in cols:
        if cell.font != bold_text or cell.row != 1 and cell.column != 1:
            cell.value = (cell.row-1) * (cell.column-1)
            if cell.value == 0:
                cell.font = bold_text


# Save the workbook.
wb.save(f'{num}_multiplication_table.xlsx')

print(f'Done. File saved as {num}_multiplication_table.xlsx.')
