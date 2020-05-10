#!/usr/bin/python3
# excel_to_csv.py - Converts all Excel files in the current working directory
# into CSV files. The program creates one CSV file per sheet.

import os
import openpyxl
import csv

for file in os.listdir('.'):
    # Skip non-Excel files, load the workbook object
    if file.endswith('.xlsx'):
        print(f'File found: {file}')
        excel_file_name = file[0:-5]
        wb = openpyxl.load_workbook(file)
        sheets = wb.sheetnames
        for sheet_name in sheets:
            sheet = wb[sheet_name]
            # Create the CSV filename from the Excel filename and sheet title
            new_file_name = excel_file_name + '_' + sheet_name + '.csv'
            output_file = open(new_file_name, 'w', newline='')
            # Create the csv.writer object for this CSV file
            output_writer = csv.writer(output_file, dialect='excel')
            # Loop through every row in the sheet.
            for row in range(1, sheet.max_row + 1):
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    if sheet.cell(row=row, column=col).value != None:
                        row_data.append(sheet.cell(row=row, column=col).value)
                # Write the row_data list to the CSV file.
                output_writer.writerow(row_data)
            output_file.close()
