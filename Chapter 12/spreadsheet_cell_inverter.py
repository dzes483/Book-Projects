#! python3
# spreadsheet_cell_inverter.py - Inverts the rows and columns in a spreadsheet.
# For example, the value at row 5, column 3 will be inverted to row 3, column
# 5. This affects all rows and columns in the spreadsheet.

import openpyxl

# Load workbooks and define sheets
source_wb = openpyxl.load_workbook('vegetables_sold.xlsx')
source_sheet = source_wb.active
target_wb = openpyxl.Workbook()
target_sheet = target_wb.active

# Use nested for loops to read the spreadsheet's data into a list of lists
max_row = source_sheet.max_row
max_col = source_sheet.max_column

print('Inverting cells...')
for row in range(1, max_row+1):
    for col in range(1, max_col+1):
        target_sheet.cell(row=col, column=row).value = source_sheet.cell(row=row, column=col).value

# Save the workbook as a copy
target_wb.save('vegetables_sold_copy.xlsx')

print('Done.')
